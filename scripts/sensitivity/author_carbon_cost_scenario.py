"""Author a CO2-price scenario directly into a clone of the master Excel's
"Carbon cost" sheet, so the scenario is visible in the source-of-truth workbook
itself rather than only in a post-prep fixture patch.

The "Carbon cost" sheet (read by excel_reader.read_carbon_costs, "new format") has one
row per ISO3/bloc in an "ISO 3-letter code_Bloc" column, with one column per year
(2020, 2021, ...) holding that country's carbon price for that year. This script:

  1. Copies --master-excel-path to --output-path byte-for-byte (every sheet other than
     "Carbon cost" is untouched).
  2. Opens the copy, finds the "Carbon cost" sheet's ISO3/Bloc column and year columns
     from the header row, and overwrites every row's year values with the same uniform
     linear-ramp schedule (params from --scenario, e.g. scenarios/co2_ramp.yaml).

The result is a normal master Excel workbook: point `steelo-data-prepare` (or
run_simulation --master-excel) at --output-path and the ramp flows through the regular
pipeline into carbon_costs.json -- no post-prep patch step, and anyone opening
--output-path in Excel sees the scenario on the "Carbon cost" tab itself.

Caveat: openpyxl doesn't round-trip every Excel feature (some chart types, certain
conditional formatting, VBA). This script only touches cell values on one sheet and
preserves everything else as openpyxl read it, but if the master workbook uses features
openpyxl can't represent, re-saving could still drop them. Spot-check --output-path
opens cleanly in Excel with its other sheets intact before relying on it for a real run.

Usage:
    uv run python -m scripts.sensitivity.author_carbon_cost_scenario \\
        --master-excel-path ~/.steelo/data_cache/master-input/master_input.xlsx \\
        --scenario scripts/sensitivity/scenarios/co2_ramp.yaml \\
        --output-path ~/.steelo/data_cache/master-input/master_input_co2_ramp.xlsx
"""

import argparse
import shutil
from pathlib import Path

import openpyxl
import yaml

ISO3_BLOC_HEADER = "ISO 3-letter code_Bloc"


def _is_year_header(value: object) -> bool:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    return value == int(value) and 1900 <= int(value) <= 2100


def _linear_ramp_schedule(
    years: list[int],
    start_price: float,
    ramp_start_year: int,
    end_price: float,
    ramp_end_year: int,
    hold_after: bool,
) -> dict[int, float]:
    if not hold_after:
        # No scenario needs this yet -- decide what "after the ramp" means once one does,
        # rather than guessing now.
        raise NotImplementedError("linear_ramp currently only supports hold_after: true")

    schedule: dict[int, float] = {}
    for year in years:
        if year <= ramp_start_year:
            price = start_price
        elif year >= ramp_end_year:
            price = end_price
        else:
            frac = (year - ramp_start_year) / (ramp_end_year - ramp_start_year)
            price = start_price + frac * (end_price - start_price)
        schedule[year] = round(price, 4)
    return schedule


def _author_carbon_cost_sheet(output_path: Path, sheet_name: str, params: dict) -> None:
    trajectory_type = params.get("type")
    if trajectory_type != "linear_ramp":
        raise NotImplementedError(f"Unsupported carbon_cost_trajectory type: {trajectory_type!r}")

    workbook = openpyxl.load_workbook(output_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {output_path}; have: {workbook.sheetnames}")
    sheet = workbook[sheet_name]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1))
    iso3_col = None
    year_cols: dict[int, int] = {}
    for cell in header_row:
        if cell.value == ISO3_BLOC_HEADER:
            iso3_col = cell.column
        elif _is_year_header(cell.value):
            year_cols[int(cell.value)] = cell.column

    if iso3_col is None:
        raise ValueError(
            f"Column {ISO3_BLOC_HEADER!r} not found in {sheet_name!r} header row of {output_path} "
            "-- is this the 'new format' Carbon cost sheet (see excel_reader.read_carbon_costs)?"
        )
    if not year_cols:
        raise ValueError(f"No year columns found in {sheet_name!r} header row of {output_path}")

    years = sorted(year_cols)
    schedule = _linear_ramp_schedule(
        years,
        start_price=params["start_price"],
        ramp_start_year=params["ramp_start_year"],
        end_price=params["end_price"],
        ramp_end_year=params["ramp_end_year"],
        hold_after=params.get("hold_after", True),
    )

    rows_written = 0
    for row in sheet.iter_rows(min_row=2):
        iso3_value = row[iso3_col - 1].value
        if not isinstance(iso3_value, str) or len(iso3_value.strip()) != 3:
            continue
        for year, col in year_cols.items():
            sheet.cell(row=row[0].row, column=col, value=schedule[year])
        rows_written += 1

    workbook.save(output_path)
    print(f"  {output_path} [{sheet_name}]: {rows_written} rows set to a uniform schedule ({years[0]}-{years[-1]})")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--master-excel-path", type=Path, required=True, help="Real master Excel workbook to clone")
    parser.add_argument("--scenario", type=Path, required=True, help="Scenario YAML, e.g. scenarios/co2_ramp.yaml")
    parser.add_argument("--output-path", type=Path, required=True, help="Where to write the authored clone")
    parser.add_argument("--sheet-name", default="Carbon cost", help="Sheet to author into (default: 'Carbon cost')")
    args = parser.parse_args()

    scenario = yaml.safe_load(args.scenario.read_text())
    if list(scenario) != ["carbon_cost_trajectory"]:
        raise ValueError(
            f"Expected exactly one 'carbon_cost_trajectory' key in {args.scenario}, got: {sorted(scenario)}"
        )

    print(f"Copying {args.master_excel_path} -> {args.output_path}")
    args.output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.master_excel_path, args.output_path)

    print(f"Authoring carbon_cost_trajectory into {args.sheet_name!r}")
    _author_carbon_cost_sheet(args.output_path, args.sheet_name, scenario["carbon_cost_trajectory"])

    print("Done.")


if __name__ == "__main__":
    main()
